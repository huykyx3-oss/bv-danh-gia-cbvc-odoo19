import base64
import io
from odoo import models, fields, api
from odoo.exceptions import UserError


class ImportEvaluationWizard(models.TransientModel):
    _name = 'bv.import.evaluation.wizard'
    _description = 'Nhập phiếu đánh giá từ Excel'

    template_id = fields.Many2one(
        'bv.evaluation.template', string='Biểu mẫu đánh giá',
        required=True, domain=[('active', '=', True), ('template_type', '=', 'monthly')])
    month = fields.Selection([
        ('1', 'Tháng 1'), ('2', 'Tháng 2'), ('3', 'Tháng 3'),
        ('4', 'Tháng 4'), ('5', 'Tháng 5'), ('6', 'Tháng 6'),
        ('7', 'Tháng 7'), ('8', 'Tháng 8'), ('9', 'Tháng 9'),
        ('10', 'Tháng 10'), ('11', 'Tháng 11'), ('12', 'Tháng 12'),
    ], string='Tháng đánh giá', required=True)
    year = fields.Integer(string='Năm', required=True,
                          default=lambda self: fields.Date.today().year)

    # Download template
    excel_template = fields.Binary(string='File mẫu Excel', readonly=True)
    excel_template_name = fields.Char(default='mau_phieu_danh_gia.xlsx')

    # Upload
    file_data = fields.Binary(string='File Excel đã điền')
    file_name = fields.Char(string='Tên file')

    # Results
    state = fields.Selection([
        ('draft', 'Chưa nhập'),
        ('done', 'Hoàn thành'),
    ], default='draft')
    result_html = fields.Html(string='Kết quả nhập liệu', readonly=True)
    success_count = fields.Integer(readonly=True)
    error_count = fields.Integer(readonly=True)

    # ──────────────────────────────────────────────────────
    # Helper: get ordered leaf criteria for the template
    # ──────────────────────────────────────────────────────
    def _get_leaf_criteria(self):
        """Return ordered list of leaf (scored) template criteria."""
        self.ensure_one()
        return self.template_id.criteria_ids.filtered(
            lambda c: c.parent_line_id
        ).sorted('sequence')

    # ──────────────────────────────────────────────────────
    # Step 1: Generate and store Excel template
    # ──────────────────────────────────────────────────────
    def action_generate_template(self):
        """Build Excel template and store it for download."""
        self.ensure_one()
        try:
            import openpyxl
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                         Border, Side, Protection)
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise UserError('Thư viện openpyxl chưa được cài đặt.')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Phiếu đánh giá'

        leaf_tcs = self._get_leaf_criteria()
        if not leaf_tcs:
            raise UserError(
                f'Biểu mẫu "{self.template_id.name}" chưa có tiêu chí lá. '
                'Hãy thêm tiêu chí con (có tiêu chí cha) trước.')

        # Style helpers
        navy = '0F2044'
        gold = 'C9A84C'
        light_blue = 'D6E4F7'
        light_gold = 'FDF8EE'
        hdr_font = Font(name='Times New Roman', bold=True, color='FFFFFF', size=11)
        hdr_fill = PatternFill('solid', fgColor=navy)
        hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        gold_fill = PatternFill('solid', fgColor=gold)
        gold_font = Font(name='Times New Roman', bold=True, color=navy, size=11)
        data_font = Font(name='Times New Roman', size=11)
        req_fill = PatternFill('solid', fgColor=light_blue)
        opt_fill = PatternFill('solid', fgColor='F5F3EE')
        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))

        # ── Row 1: Title ──
        month_label = dict(self._fields['month'].selection).get(self.month, '')
        ws.merge_cells('A1:A1')
        title_text = (
            f'PHIẾU ĐÁNH GIÁ HẰNG THÁNG — {month_label.upper()} NĂM {self.year}\n'
            f'Biểu mẫu: {self.template_id.name}\n'
            '(*) Cột bắt buộc | Điểm tiêu chí: nhập số trong phạm vi 0 → điểm tối đa | '
            '% Nhiệm vụ: nhập số 0–100'
        )
        last_col_idx = 6 + len(leaf_tcs) + 5  # fixed cols + criteria + task cols
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=last_col_idx)
        c = ws.cell(row=1, column=1, value=title_text)
        c.font = Font(name='Times New Roman', bold=True, color='FFFFFF', size=12)
        c.fill = PatternFill('solid', fgColor=navy)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 52

        # ── Row 2: Section headers ──
        # Fixed cols: STT | Mã NV* | Họ và tên* | Khoa/Phòng | Tháng* | Năm*
        fixed_headers = [
            ('STT', False), ('Mã NV', True),
            ('Họ và tên', True), ('Khoa/Phòng', False),
            ('Tháng', True), ('Năm', True),
        ]
        col = 1
        for label, required in fixed_headers:
            display = f'{label}{"*" if required else ""}'
            c = ws.cell(row=2, column=col, value=display)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = hdr_align
            c.border = thin
            col += 1

        # Criteria columns (with max score hint)
        for tc in leaf_tcs:
            group_name = tc.parent_line_id.name[:20] + '...' \
                if len(tc.parent_line_id.name) > 20 else tc.parent_line_id.name
            header = f'{group_name}\n{tc.name[:30]}\n(max: {tc.max_score}đ)*'
            c = ws.cell(row=2, column=col, value=header)
            c.font = gold_font
            c.fill = gold_fill
            c.alignment = hdr_align
            c.border = thin
            ws.column_dimensions[get_column_letter(col)].width = 22
            col += 1

        # Task % columns
        task_cols = [
            ('% Số lượng\n(0-100)*', True),
            ('% Chất lượng\n(0-100)*', True),
            ('% Tiến độ\n(0-100)*', True),
            ('% Kết quả\nlĩnh vực*\n(Lãnh đạo)', False),
            ('% Tổ chức\ntriển khai*\n(Lãnh đạo)', False),
            ('% Đoàn kết\n(Lãnh đạo)', False),
        ]
        for label, required in task_cols:
            c = ws.cell(row=2, column=col, value=label)
            c.font = hdr_font
            c.fill = PatternFill('solid', fgColor='2D5A27' if required else '4A8C43')
            c.alignment = hdr_align
            c.border = thin
            ws.column_dimensions[get_column_letter(col)].width = 14
            col += 1

        # Optional text cols
        for label in ['Ưu điểm (tùy chọn)', 'Hạn chế (tùy chọn)']:
            c = ws.cell(row=2, column=col, value=label)
            c.font = Font(name='Times New Roman', bold=True, color='1E1E1E', size=10)
            c.fill = PatternFill('solid', fgColor='ECEAE4')
            c.alignment = hdr_align
            c.border = thin
            ws.column_dimensions[get_column_letter(col)].width = 22
            col += 1

        ws.row_dimensions[2].height = 56

        # Fixed col widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 26
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 8

        # ── Rows 3–52: Sample data rows (50 employees) ──
        for row_idx in range(3, 53):
            i = row_idx - 2
            # STT
            c = ws.cell(row=row_idx, column=1, value=i)
            c.font = data_font
            c.alignment = Alignment(horizontal='center')
            c.border = thin
            # Month/Year defaults
            ws.cell(row=row_idx, column=5, value=int(self.month))
            ws.cell(row=row_idx, column=6, value=self.year)
            for ci in range(1, last_col_idx + 1):
                ws.cell(row=row_idx, column=ci).border = thin
                ws.cell(row=row_idx, column=ci).font = data_font
                if ci in (5, 6):
                    ws.cell(row=row_idx, column=ci).alignment = Alignment(horizontal='center')
            # Color bands
            fill = req_fill if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
            for ci in range(2, last_col_idx + 1):
                ws.cell(row=row_idx, column=ci).fill = fill

        # Freeze panes
        ws.freeze_panes = 'G3'

        # ── Sheet 2: Instructions ──
        ws2 = wb.create_sheet(title='Hướng dẫn')
        instructions = [
            ('HƯỚNG DẪN NHẬP LIỆU', True),
            ('', False),
            ('1. Điền thông tin vào sheet "Phiếu đánh giá" (không xóa dòng tiêu đề)', False),
            ('2. Cột Mã NV: mã nhân viên trong hệ thống (tìm trong module Nhân viên)', False),
            ('3. Họ và tên: đúng tên trong hệ thống (dùng để tìm nếu không có mã)', False),
            ('4. Khoa/Phòng: không bắt buộc nếu đã có Mã NV hoặc Họ tên chính xác', False),
            ('5. Điểm tiêu chí: nhập số thực, từ 0 đến điểm tối đa ghi trong tiêu đề cột', False),
            ('6. % Nhiệm vụ: nhập số nguyên hoặc thực, từ 0 đến 100', False),
            ('7. Cột lãnh đạo (% Kết quả lĩnh vực, % Tổ chức, % Đoàn kết): '
             'để trống nếu không phải lãnh đạo', False),
            ('8. Sau khi điền xong, lưu file và tải lên ô "File Excel đã điền"', False),
            ('9. Nhấn "Nhập dữ liệu" để tạo phiếu tự động', False),
            ('', False),
            ('LƯU Ý:', True),
            ('- Không thay đổi tên sheet "Phiếu đánh giá"', False),
            ('- Không thay đổi hàng tiêu đề (hàng 2)', False),
            ('- Phiếu đã tồn tại (cùng nhân viên + tháng + năm) sẽ bị bỏ qua', False),
            ('- Hệ thống tự động tìm nhân viên theo Mã NV, nếu không có thì tìm theo Họ tên', False),
        ]
        for i, (text, bold) in enumerate(instructions, 1):
            c = ws2.cell(row=i, column=1, value=text)
            c.font = Font(name='Times New Roman', bold=bold, size=12,
                          color=navy if bold else '1E1E1E')
            if bold:
                c.fill = PatternFill('solid', fgColor=gold)
        ws2.column_dimensions['A'].width = 90

        # Save
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        file_b64 = base64.b64encode(buf.getvalue())
        fname = (f'mau_phieu_danh_gia_{self.template_id.code or self.template_id.name}'
                 f'_T{self.month}_{self.year}.xlsx').replace(' ', '_')
        self.write({'excel_template': file_b64, 'excel_template_name': fname})
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    # ──────────────────────────────────────────────────────
    # Step 2: Import uploaded Excel
    # ──────────────────────────────────────────────────────
    def action_import(self):
        self.ensure_one()
        if not self.file_data:
            raise UserError('Vui lòng tải lên file Excel trước khi nhập.')

        try:
            import openpyxl
        except ImportError:
            raise UserError('Thư viện openpyxl chưa được cài đặt.')

        raw = base64.b64decode(self.file_data)
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        except Exception as e:
            raise UserError(f'Không đọc được file Excel: {e}')

        if 'Phiếu đánh giá' not in wb.sheetnames:
            raise UserError('File Excel không đúng định dạng. '
                            'Vui lòng tải file mẫu và điền vào đó.')

        ws = wb['Phiếu đánh giá']
        leaf_tcs = self._get_leaf_criteria()

        # Determine column mapping from header row (row 2)
        header_row = [ws.cell(row=2, column=c).value or ''
                      for c in range(1, ws.max_column + 1)]

        # Fixed columns: STT(0), Mã NV(1), Họ và tên(2), Khoa/Phòng(3), Tháng(4), Năm(5)
        # Criteria columns: 6..6+len(leaf_tcs)-1
        # Task columns: quantity, quality, progress, field_result, organization, team_cohesion
        criteria_col_start = 6  # 0-indexed
        task_col_start = criteria_col_start + len(leaf_tcs)
        task_keys = ['pct_quantity', 'pct_quality', 'pct_progress',
                     'pct_field_result', 'pct_organization', 'pct_team_cohesion']

        Employee = self.env['hr.employee']
        Evaluation = self.env['bv.monthly.evaluation']
        CriteriaLine = self.env['bv.evaluation.criteria.line']

        success_rows, error_rows = [], []

        for row_idx in range(3, ws.max_row + 1):
            row = [ws.cell(row=row_idx, column=c).value
                   for c in range(1, ws.max_column + 1)]

            # Skip empty rows
            if not any(v for v in row[1:6] if v):
                continue

            emp_code = str(row[1] or '').strip()
            emp_name = str(row[2] or '').strip()
            dept_name = str(row[3] or '').strip()

            # Try month/year from row, else use wizard values
            try:
                month_val = str(int(row[4])) if row[4] else self.month
                year_val = int(row[5]) if row[5] else self.year
            except (TypeError, ValueError):
                month_val = self.month
                year_val = self.year

            # Find employee
            employee = False
            if emp_code:
                employee = Employee.search([
                    '|', ('barcode', '=', emp_code),
                    ('employee_id', '=', emp_code),
                ], limit=1)
            if not employee and emp_name:
                domain = [('name', 'ilike', emp_name)]
                if dept_name:
                    domain += [('department_id.name', 'ilike', dept_name)]
                employee = Employee.search(domain, limit=1)
            if not employee and emp_name:
                # Broader search without department
                employee = Employee.search([('name', 'ilike', emp_name)], limit=1)

            if not employee:
                error_rows.append({
                    'row': row_idx, 'name': emp_name or emp_code,
                    'reason': 'Không tìm thấy nhân viên trong hệ thống',
                })
                continue

            # Check duplicate
            existing = Evaluation.search([
                ('employee_id', '=', employee.id),
                ('month', '=', month_val),
                ('year', '=', year_val),
            ], limit=1)
            if existing:
                error_rows.append({
                    'row': row_idx, 'name': employee.name,
                    'reason': f'Phiếu tháng {month_val}/{year_val} đã tồn tại (ID: {existing.id})',
                })
                continue

            # Parse criteria scores
            criteria_scores = {}
            for i, tc in enumerate(leaf_tcs):
                col_idx = criteria_col_start + i
                raw_val = row[col_idx] if col_idx < len(row) else None
                try:
                    score = float(raw_val) if raw_val is not None and raw_val != '' else 0.0
                    # Clamp to [0, max_score]
                    score = max(0.0, min(score, tc.max_score))
                except (TypeError, ValueError):
                    score = 0.0
                if tc.synced_criteria_id:
                    criteria_scores[tc.synced_criteria_id.id] = score

            # Parse task percentages
            task_vals = {}
            for j, key in enumerate(task_keys):
                col_idx = task_col_start + j
                raw_val = row[col_idx] if col_idx < len(row) else None
                try:
                    pct = float(raw_val) if raw_val is not None and raw_val != '' else 0.0
                    pct = max(0.0, min(pct, 100.0))
                except (TypeError, ValueError):
                    pct = 0.0
                task_vals[key] = pct

            # Strengths / weaknesses
            strengths_col = task_col_start + len(task_keys)
            weakness_col = strengths_col + 1
            strengths = str(row[strengths_col] or '') if strengths_col < len(row) else ''
            weaknesses = str(row[weakness_col] or '') if weakness_col < len(row) else ''

            try:
                is_manager = bool(
                    task_vals.get('pct_field_result') or
                    task_vals.get('pct_organization') or
                    task_vals.get('pct_team_cohesion')
                )
                eval_vals = {
                    'employee_id': employee.id,
                    'department_id': employee.department_id.id,
                    'job_id': employee.job_id.id if employee.job_id else False,
                    'month': month_val,
                    'year': year_val,
                    'template_id': self.template_id.id,
                    'is_manager': is_manager,
                    'strengths': strengths,
                    'weaknesses': weaknesses,
                    **task_vals,
                }
                new_eval = Evaluation.create(eval_vals)

                # Update criteria scores
                for line in new_eval.criteria_line_ids:
                    if line.criteria_id.id in criteria_scores:
                        line.self_score = criteria_scores[line.criteria_id.id]

                success_rows.append({
                    'row': row_idx, 'name': employee.name,
                    'month': month_val, 'year': year_val,
                    'eval_id': new_eval.id,
                })
            except Exception as e:
                error_rows.append({
                    'row': row_idx, 'name': employee.name,
                    'reason': str(e)[:200],
                })

        # Build result HTML
        html = self._build_result_html(success_rows, error_rows)
        self.write({
            'state': 'done',
            'result_html': html,
            'success_count': len(success_rows),
            'error_count': len(error_rows),
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _build_result_html(self, success_rows, error_rows):
        lines = ['<div style="font-family:\'Times New Roman\',serif;">']
        total = len(success_rows) + len(error_rows)
        error_badge = (
            ' — <span style="color:#b03030;font-weight:700">'
            + str(len(error_rows)) + ' lỗi</span>'
        ) if error_rows else ''
        lines.append(
            '<div style="padding:12px 16px;background:#e8f5ed;border-left:4px solid #2d7a4f;'
            'border-radius:6px;margin-bottom:16px;">'
            '<strong>Kết quả nhập liệu:</strong> '
            'Tổng ' + str(total) + ' dòng — '
            '<span style="color:#2d7a4f;font-weight:700">'
            + str(len(success_rows)) + ' thành công</span>'
            + error_badge +
            '</div>'
        )
        if success_rows:
            lines.append(
                '<table style="width:100%;border-collapse:collapse;margin-bottom:16px;">'
                '<thead><tr>'
                '<th style="background:#0f2044;color:white;padding:8px;text-align:left;">Dòng</th>'
                '<th style="background:#0f2044;color:white;padding:8px;text-align:left;">Nhân viên</th>'
                '<th style="background:#0f2044;color:white;padding:8px;text-align:left;">Kỳ</th>'
                '<th style="background:#0f2044;color:white;padding:8px;text-align:left;">Trạng thái</th>'
                '</tr></thead><tbody>'
            )
            for i, r in enumerate(success_rows):
                bg = '#f5f3ee' if i % 2 else '#ffffff'
                lines.append(
                    f'<tr style="background:{bg};">'
                    f'<td style="padding:6px 8px;border-bottom:1px solid #eceae4;">{r["row"]}</td>'
                    f'<td style="padding:6px 8px;border-bottom:1px solid #eceae4;">{r["name"]}</td>'
                    f'<td style="padding:6px 8px;border-bottom:1px solid #eceae4;">'
                    f'Tháng {r["month"]}/{r["year"]}</td>'
                    f'<td style="padding:6px 8px;border-bottom:1px solid #eceae4;color:#2d7a4f;">'
                    f'✔ Đã tạo (ID: {r["eval_id"]})</td>'
                    f'</tr>'
                )
            lines.append('</tbody></table>')
        if error_rows:
            lines.append(
                '<div style="background:#fff3cd;border-left:4px solid #ffc107;'
                'padding:10px 16px;border-radius:6px;margin-bottom:8px;">'
                '<strong>Các dòng lỗi:</strong></div>'
                '<table style="width:100%;border-collapse:collapse;">'
                '<thead><tr>'
                '<th style="background:#b03030;color:white;padding:8px;">Dòng</th>'
                '<th style="background:#b03030;color:white;padding:8px;">Nhân viên</th>'
                '<th style="background:#b03030;color:white;padding:8px;">Lý do</th>'
                '</tr></thead><tbody>'
            )
            for i, r in enumerate(error_rows):
                bg = '#fdf0f0' if i % 2 else '#fff8f8'
                lines.append(
                    f'<tr style="background:{bg};">'
                    f'<td style="padding:6px 8px;border-bottom:1px solid #f8d7da;">{r["row"]}</td>'
                    f'<td style="padding:6px 8px;border-bottom:1px solid #f8d7da;">{r["name"]}</td>'
                    f'<td style="padding:6px 8px;border-bottom:1px solid #f8d7da;color:#b03030;">'
                    f'{r["reason"]}</td>'
                    f'</tr>'
                )
            lines.append('</tbody></table>')
        lines.append('</div>')
        return ''.join(lines)

    def action_open_evaluations(self):
        """Open the list of created evaluations after import."""
        return {
            'type': 'ir.actions.act_window',
            'name': 'Phiếu đánh giá vừa nhập',
            'res_model': 'bv.monthly.evaluation',
            'view_mode': 'list,form',
            'domain': [('template_id', '=', self.template_id.id),
                       ('month', '=', self.month),
                       ('year', '=', self.year)],
            'target': 'current',
        }
